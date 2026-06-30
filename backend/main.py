import os
import sys
import openpyxl
import pandas as pd

sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from fastapi import FastAPI, HTTPException, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy.orm import Session
from typing import Optional, List

from database import get_db, engine
import models

from fastapi.responses import FileResponse, StreamingResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io

# Crear tablas
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="API GAC - Clínica San Rafael")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==========================================
# MODELOS DE DATOS 
# ==========================================
class ActivoInput(BaseModel):
    codigo_activo: str
    placa: str
    nombre: str
    serie: Optional[str] = "N/A"
    ubicacion_origen: Optional[str] = "N/A"
    ubicacion_destino: Optional[str] = "N/A"
    responsable_origen: Optional[str] = "N/A"
    responsable_destino: Optional[str] = "N/A"
    centro_costos_origen: Optional[int] = 0
    centro_costos_destino: Optional[int] = 0
    porcentaje: Optional[float] = 100.0
    secuencia: Optional[str] = "N/A"
    enlace: Optional[str] = ""
    archivo_plano: Optional[str] = ""

class MovimientoInput(BaseModel):
    codigo_barras: str
    id_usuario: int
    nueva_ubicacion: str
    observaciones: Optional[str] = None

class ItemActa(BaseModel):
    cant: int
    desc: str
    placa: str
    marca: str
    modelo: str
    serie: str

class DatosActa(BaseModel):
    dia: str
    mes: str
    anio: str
    concepto: str
    numero: str
    tipo: str
    proveedor: str
    factura: str
    resp_nombre: str
    resp_doc: str
    resp_cargo: str
    items: List[ItemActa]

# ==========================================
# ENDPOINTS / RUTAS
# ==========================================
@app.get("/")
def ruta_principal():
    return {"mensaje": "¡Backend GAC funcionando!"}

@app.post("/activos")
def registrar_nuevo_activo(activo: ActivoInput, db: Session = Depends(get_db)):
    existe = db.query(models.ActivoDB).filter(models.ActivoDB.codigo_barras == activo.placa).first()
    if existe:
        raise HTTPException(status_code=400, detail="Error: Esta placa ya existe.")
    
    nuevo_equipo = models.ActivoDB(
        codigo_barras=activo.placa,
        codigo_activo=activo.codigo_activo,
        nombre=activo.nombre,
        serie=activo.serie,
        ubicacion_origen=activo.ubicacion_origen,
        ubicacion_destino=activo.ubicacion_destino,
        responsable_origen=activo.responsable_origen,
        responsable_destino=activo.responsable_destino,
        centro_costos_origen=activo.centro_costos_origen,
        centro_costos_destino=activo.centro_costos_destino,
        porcentaje=activo.porcentaje,
        secuencia=activo.secuencia,
        enlace=activo.enlace,
        archivo_plano=activo.archivo_plano,
        estado="Operativo"
    )
    db.add(nuevo_equipo)
    db.commit()
    return {"estado": "Éxito", "mensaje": "¡Activo registrado contablemente!"}

@app.get("/activos/{codigo_barras}")
def consultar_activo(codigo_barras: str, db: Session = Depends(get_db)):
    activo = db.query(models.ActivoDB).filter(models.ActivoDB.codigo_barras == codigo_barras).first()
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no registrado")
    return activo

@app.post("/movimientos")
def registrar_movimiento(datos: MovimientoInput, db: Session = Depends(get_db)):
    activo = db.query(models.ActivoDB).filter(models.ActivoDB.codigo_barras == datos.codigo_barras).first()
    if not activo:
        raise HTTPException(status_code=404, detail="Activo no existe")
    
    ubicacion_anterior = activo.ubicacion_destino
    try:
        activo.ubicacion_destino = datos.nueva_ubicacion
        nuevo_traslado = models.MovimientoDB(
            codigo_barras=datos.codigo_barras,
            id_usuario=int(datos.id_usuario),
            ubicacion_anterior=ubicacion_anterior,
            nueva_ubicacion=datos.nueva_ubicacion,
            observaciones=datos.observaciones
        )
        db.add(nuevo_traslado)
        db.commit()
        return {"estado": "Éxito", "mensaje": "Traslado guardado."}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail="Error al registrar el traslado.")

@app.get("/historial")
def ver_historial(db: Session = Depends(get_db)):
    registros = db.query(models.MovimientoDB).all()
    return {"total_movimientos": len(registros), "registros": registros}

@app.post("/api/actas/generar")
async def generar_acta_excel(datos: DatosActa):
    ruta_plantilla = r"C:\Users\57318\Documents\Proyecto Inventario Activos Fijos\gac-sistema\backend\plantilla_acta.xlsx"
    wb = openpyxl.load_workbook(ruta_plantilla)
    nombre_primera_hoja = wb.sheetnames[0]
    ws = wb[nombre_primera_hoja]
    
    ws["C4"] = datos.dia
    ws["D4"] = datos.mes
    ws["E4"] = datos.anio
    ws["H5"] = datos.numero
    
    celdas_concepto = {
        'Donación': 'C10', 'Comodato': 'D10', 'Préstamo Tercero': 'E10',
        'Demostración': 'F10', 'Compra': 'G10', 'Arrendamiento': 'H10', 'Leasing': 'I10'
    }
    if datos.concepto in celdas_concepto:
        ws[celdas_concepto[datos.concepto]] = "X"

    ws["B30"] = datos.proveedor
    ws["E30"] = datos.factura
    ws["B35"] = datos.resp_nombre
    ws["B36"] = datos.resp_doc
    ws["B37"] = datos.resp_cargo

    fila_inicio = 20
    for i, item in enumerate(datos.items):
        fila_actual = fila_inicio + i
        ws[f"A{fila_actual}"] = item.cant
        ws[f"B{fila_actual}"] = item.desc
        ws[f"C{fila_actual}"] = item.placa
        ws[f"D{fila_actual}"] = item.marca
        ws[f"E{fila_actual}"] = item.modelo
        ws[f"F{fila_actual}"] = item.serie

    nombre_responsable = datos.resp_nombre if datos.resp_nombre else "General"
    nombre_archivo = f"Acta_Entrada_{nombre_responsable.replace(' ', '_')}.xlsx"
    ruta_guardado = os.path.join(r"C:\Users\57318\Documents\Proyecto Inventario Activos Fijos\gac-sistema\backend", nombre_archivo)
    
    wb.save(ruta_guardado)

    return FileResponse(
        path=ruta_guardado, 
        filename=nombre_archivo, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/inventario-completo")
def obtener_inventario():
    ruta_csv = r"C:\Users\57318\Documents\Proyecto Inventario Activos Fijos\gac-sistema\backend\BASE DE DATOS SISTEMA GAC.xlsx - Activos por Código.csv"
    try:
        df = pd.read_csv(ruta_csv)
        df = df.fillna("") 
        return df.to_dict(orient="records")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# MONTAR ARCHIVOS ESTÁTICOS
# ==========================================


# ... (Todo tu código anterior de main.py, como ActivoInput, MovimientoInput, etc.)

@app.post("/api/actas/generar_pdf")
async def generar_acta_pdf(datos: DatosActa):
    # Creamos un buffer en memoria para el PDF
    buffer = io.BytesIO()
    
    # Configuramos el documento (Tamaño carta, márgenes pequeños)
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            rightMargin=30, leftMargin=30,
                            topMargin=30, bottomMargin=30)
    
    elementos = []
    estilos = getSampleStyleSheet()
    estilo_titulo = estilos['Heading1']
    estilo_titulo.alignment = 1 # Centrado
    estilo_normal = estilos['Normal']

    # --- 1. ENCABEZADO (Membete) ---
    # Intentamos cargar el logo (asegúrate de tener 'logo_clinica.png' en la carpeta)
    ruta_logo = "logo_clinica.png" # <--- CAMBIA ESTO SI TU LOGO ESTÁ EN OTRA RUTA
    try:
        logo = RLImage(ruta_logo, width=1.5*inch, height=1.5*inch)
    except:
        logo = Paragraph("<b>[LOGO CLINICA]</b>", estilo_normal) # Placeholder si no hay logo

    datos_encabezado = [
        [logo, "CLINICA SAN RAFAEL POPAYAN\nGESTION DE ACTIVOS", f"FECHA: {datos.dia}/{datos.mes}/{datos.anio}\nACTA N°: {datos.numero}"]
    ]
    tabla_encabezado = Table(datos_encabezado, colWidths=[2*inch, 3*inch, 2.5*inch])
    tabla_encabezado.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ]))
    elementos.append(tabla_encabezado)
    elementos.append(Spacer(1, 12))

    # --- 2. TÍTULO DEL DOCUMENTO ---
    elementos.append(Paragraph("<b>ACTA DE INGRESO Y RECEPCIÓN DE ACTIVOS FIJOS</b>", estilo_titulo))
    elementos.append(Spacer(1, 12))

    # --- 3. INFORMACIÓN GENERAL (Concepto, Tipo, Proveedor) ---
    datos_info = [
        ["CONCEPTO:", datos.concepto, "TIPO DE ACTIVO:", datos.tipo],
        ["PROVEEDOR:", datos.proveedor, "FACTURA N°:", datos.factura]
    ]
    tabla_info = Table(datos_info, colWidths=[1.5*inch, 2.25*inch, 1.5*inch, 2.25*inch])
    tabla_info.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.lightgrey),
        ('BACKGROUND', (2,0), (2,-1), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('PADDING', (0,0), (-1,-1), 6)
    ]))
    elementos.append(tabla_info)
    elementos.append(Spacer(1, 12))

    # --- 4. TABLA DE ACTIVOS (Dinámica) ---
    # Definimos las cabeceras de la tabla
    datos_activos = [["CANT", "DESCRIPCIÓN", "PLACA / CÓDIGO", "MARCA", "MODELO", "SERIE"]]
    
    # Agregamos las filas según los items que llegaron del formulario
    for item in datos.items:
        datos_activos.append([str(item.cant), item.desc, item.placa, item.marca, item.modelo, item.serie])

    # Calculamos anchos para que encajen bien
    anchos_columnas = [0.6*inch, 2.5*inch, 1.4*inch, 1.0*inch, 1.0*inch, 1.0*inch]
    tabla_activos = Table(datos_activos, colWidths=anchos_columnas, repeatRows=1)
    
    tabla_activos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.Color(0.2, 0.4, 0.8)), # Azul corporativo
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ALIGN', (0,1), (0,-1), 'CENTER'), # Centrar cantidades
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elementos.append(tabla_activos)
    elementos.append(Spacer(1, 30))

    # --- 5. FIRMAS ---
    datos_firmas = [
        ["ENTREGA:", "RECIBE:"],
        ["_________________________", "_________________________"],
        ["Nombre: ________________", f"Nombre: {datos.resp_nombre}"],
        ["Cargo: _________________", f"Cargo: {datos.resp_cargo}"],
        ["Documento: ______________", f"Documento: {datos.resp_doc}"]
    ]
    tabla_firmas = Table(datos_firmas, colWidths=[3.75*inch, 3.75*inch])
    tabla_firmas.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
        ('BOTTOMPADDING', (0,1), (-1,1), 20), # Espacio para la firma
    ]))
    elementos.append(tabla_firmas)

    # --- 6. GENERAR EL PDF ---
    doc.build(elementos)
    
    # Preparamos la respuesta para descargar
    buffer.seek(0)
    nombre_archivo = f"Acta_Ingreso_{datos.resp_nombre.replace(' ', '_')}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )

# ... (El resto de tu main.py, ej: app.mount)
app.mount("/", StaticFiles(directory="../frontend", html=True), name="static")